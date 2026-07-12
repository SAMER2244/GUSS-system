"""
migrate_legacy_data.py — ترحيل بيانات الإكسل التاريخية إلى Supabase
=====================================================================

الاستخدام:
    # عرض تقرير dry-run (لا كتابة):
    python migrate_legacy_data.py source.xlsx

    # تنفيذ فعلي بعد المراجعة:
    python migrate_legacy_data.py source.xlsx --commit

    # تحديد ملف المطابقة صراحةً:
    python migrate_legacy_data.py source.xlsx --commit --mapping office_name_mapping.json

ملاحظة أمنية: لا يُنفَّذ أي إدخال فعلي إلا بتمرير --commit صراحةً.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import openpyxl
from tqdm import tqdm
import config

# ─── المسارات ────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
MIGRATION_DIR = BASE_DIR / "data" / "migration"
DEFAULT_MAPPING_FILE = BASE_DIR / "office_name_mapping.json"

# ─── ثوابت البنية ─────────────────────────────────────────────────────────────
# Sheet 2 ("ردود النموذج 2"):
#   col 1 : طابع زمني
#   col 2 : اسم المكتب
#   col 3 : المسؤول عن التعبئة
#   col 4 : رقم الهاتف
#   col 5 : ملف الخطة (plan_file URL) — خاص بـ Sheet2
#   cols 6-115 : 11 بلوك مهام (10 أعمدة لكل بلوك)
#   col 116: التحديات العامة
#   col 117: ملاحظات إضافية
#   col 118: الشهر (عمود موجود فقط في Sheet2 لكن كان فارغاً)
#
# Sheet 1 ("ردود النموذج 1"):
#   col 1 : طابع زمني
#   col 2 : اسم المكتب
#   col 3 : المسؤول عن التعبئة
#   col 4 : رقم الهاتف
#   cols 5-114 : 11 بلوك مهام (10 أعمدة لكل بلوك) — بدون عمود خطة في البداية
#   col 115: ملف الخطة
#   col 116: التحديات العامة
#   col 117: ملاحظات إضافية
#   (لا يوجد col 118)

SHEET2_NAME = "ردود النموذج 2"
SHEET1_NAME = "ردود النموذج 1"

# بلوك المهام: 10 أعمدة لكل بلوك بترتيب ثابت بغض النظر عن الاسم الحرفي
TASK_BLOCK_COLS = 10
MAX_TASK_BLOCKS = 11

# موضع كل حقل داخل البلوك (0-based offset داخل البلوك):
BLOCK_OFFSET = {
    "manager_name": 0,
    "manager_phone": 1,
    "task_name": 2,
    "task_description": 3,
    "task_type": 4,
    "execution_mechanism": 5,
    "task_status": 6,
    "issues": 7,
    "file_attach": 8,
    "add_more": 9,
}

# نمط الاستبعاد لصفوف الاختبار
TEST_PATTERNS = [
    "تجريبي", "اختبار", "test", "demo", "بيانات تجريبية",
    "dummy", "fake", "sample", "نموذج تجريبي",
]

ARABIC_MONTHS = {
    "يناير": 1, "فبراير": 2, "مارس": 3, "أبريل": 4, "إبريل": 4,
    "مايو": 5, "يونيو": 6, "يوليو": 7, "أغسطس": 8,
    "سبتمبر": 9, "أكتوبر": 10, "نوفمبر": 11, "ديسمبر": 12,
    "كانون الثاني": 1, "شباط": 2, "آذار": 3, "نيسان": 4,
    "أيار": 5, "حزيران": 6, "تموز": 7, "آب": 8,
    "أيلول": 9, "تشرين الأول": 10, "تشرين الثاني": 11, "كانون الأول": 12,
}


# ─── مساعدات تنظيف البيانات ──────────────────────────────────────────────────

def clean_phone(raw) -> str:
    """ينظّف رقم الهاتف: يزيل Unicode خفي، يوحّد الفواصل."""
    if raw is None:
        return ""
    s = str(raw)
    # إزالة unicode formatting characters
    s = "".join(
        c for c in s
        if not unicodedata.category(c).startswith("C")
        and unicodedata.category(c) != "Zs"
        or c == " "
    )
    s = s.strip()
    # توحيد الفواصل بين أرقام متعددة
    s = re.sub(r"[،,;/]+", " / ", s)
    # إزالة مسافات متعددة
    s = re.sub(r"\s{2,}", " ", s)
    return s.strip()


def clean_text(raw) -> str:
    """تنظيف نصي عام: إزالة مسافات زائدة وRTL markers."""
    if raw is None:
        return ""
    s = str(raw).strip()
    # إزالة RTL/LTR marks
    s = s.replace("\u200f", "").replace("\u200e", "").replace("\u200b", "")
    return s.strip()


def is_test_row(row_dict: dict) -> bool:
    """يكتشف صفوف الاختبار بناءً على وجود أنماط الاختبار ككلمات مستقلة أو مطابقة كاملة."""
    texts = [
        row_dict.get("office_raw", ""),
        row_dict.get("submitter_name", ""),
        row_dict.get("general_challenges", ""),
        row_dict.get("additional_notes", ""),
    ]
    for t in row_dict.get("tasks", []):
        texts.append(t.get("task_name", ""))
        texts.append(t.get("task_description", ""))
        texts.append(t.get("issues", ""))

    def contains_word(text: str) -> bool:
        if not text:
            return False
        # تنظيف النص من علامات الترقيم وتبسيط الفراغات
        cleaned = re.sub(r"[^\w\s]", " ", text.lower())
        words = cleaned.split()
        for w in words:
            if w in ["تجريبي", "اختبار", "test", "demo", "dummy", "fake", "sample"]:
                return True
        # مطابقة عبارات كاملة
        lower_text = text.lower()
        if "بيانات تجريبية" in lower_text or "نموذج تجريبي" in lower_text:
            return True
        return False

    for txt in texts:
        if contains_word(txt):
            return True

    return False


def extract_month_from_text(text: str) -> tuple[int | None, int | None]:
    """يستخرج الشهر والسنة من نص حر (مثل 'شهر مارس 2026')."""
    if not text:
        return None, None
    for ar_name, num in ARABIC_MONTHS.items():
        if ar_name in text:
            # محاولة استخراج السنة
            year_match = re.search(r"(202\d|203\d)", text)
            year = int(year_match.group(1)) if year_match else None
            return num, year
    # رقمي مثل "3/2026" أو "2026-03"
    m = re.search(r"(\d{1,2})[/-](\d{4})", text)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def similarity(a: str, b: str) -> float:
    """نسبة التشابه بين نصين."""
    return SequenceMatcher(None, a.strip(), b.strip()).ratio()


# ─── تفكيك الصف من الشيت ─────────────────────────────────────────────────────

def parse_row(ws, row: int, sheet_type: str) -> dict:
    """
    يفكّك صفاً كاملاً من الشيت ويُعيد dict موحّداً.

    sheet_type: 'sheet1' أو 'sheet2'
    """
    def cell(c): return ws.cell(row, c).value

    ts = cell(1)
    office_raw = clean_text(cell(2))
    submitter = clean_text(cell(3))
    phone = clean_phone(cell(4))

    if sheet_type == "sheet2":
        plan_file = clean_text(cell(5))
        task_start_col = 6       # أول بلوك مهمة يبدأ من col 6
        challenges_col = 116
        notes_col = 117
        month_col = 118
    else:
        # sheet1: لا يوجد عمود خطة في البداية
        plan_file = clean_text(cell(115))
        task_start_col = 5       # أول بلوك مهمة يبدأ من col 5
        challenges_col = 116
        notes_col = 117
        month_col = None

    # ─── استخراج الشهر/السنة ─────────────────────────────────────────────────
    month_from_col = None
    year_from_col = None
    month_source = "timestamp"   # المصدر المستخدم

    if month_col:
        month_text = clean_text(cell(month_col))
        if month_text:
            m, y = extract_month_from_text(month_text)
            if m:
                month_from_col, year_from_col = m, y
                month_source = "month_column"

    if month_from_col is None and ts and hasattr(ts, "month"):
        month_from_col = ts.month
        year_from_col = ts.year
        month_source = "timestamp"

    # ─── استخراج بلوكات المهام بالموقع ──────────────────────────────────────
    tasks = []
    for block_idx in range(MAX_TASK_BLOCKS):
        base = task_start_col + block_idx * TASK_BLOCK_COLS
        manager_name = clean_text(cell(base + BLOCK_OFFSET["manager_name"]))
        manager_phone = clean_phone(cell(base + BLOCK_OFFSET["manager_phone"]))
        task_name = clean_text(cell(base + BLOCK_OFFSET["task_name"]))
        task_description = clean_text(cell(base + BLOCK_OFFSET["task_description"]))
        task_type = clean_text(cell(base + BLOCK_OFFSET["task_type"]))
        execution_mechanism = clean_text(cell(base + BLOCK_OFFSET["execution_mechanism"]))
        task_status = clean_text(cell(base + BLOCK_OFFSET["task_status"]))
        issues = clean_text(cell(base + BLOCK_OFFSET["issues"]))

        if not task_name:
            continue   # بلوك فارغ

        tasks.append({
            "manager_name": manager_name,
            "manager_phone": manager_phone,
            "task_name": task_name,
            "task_description": task_description,
            "task_type": task_type,
            "execution_mechanism": execution_mechanism,
            "task_status": task_status,
            "issues": issues,
        })

    return {
        "row": row,
        "sheet": sheet_type,
        "timestamp": ts,
        "office_raw": office_raw,
        "submitter_name": submitter,
        "submitter_phone": phone,
        "plan_file": plan_file,
        "month": month_from_col,
        "year": year_from_col,
        "month_source": month_source,
        "general_challenges": clean_text(cell(challenges_col)),
        "additional_notes": clean_text(cell(notes_col)),
        "tasks": tasks,
    }


# ─── تحميل بيانات الإكسل ─────────────────────────────────────────────────────

def load_excel(path: Path) -> tuple[list[dict], list[dict]]:
    """
    يُحمّل الشيتين ويُعيد قائمتين: (rows_accepted, rows_excluded).
    كل عنصر dict يحمل حقل 'exclude_reason' إذا استُبعد.
    """
    wb = openpyxl.load_workbook(str(path))
    accepted = []
    excluded = []

    for sheet_name, sheet_type in [
        (SHEET2_NAME, "sheet2"),
        (SHEET1_NAME, "sheet1"),
    ]:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  الشيت '{sheet_name}' غير موجود في الملف — تجاهل.")
            continue

        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            ts = ws.cell(r, 1).value
            office = ws.cell(r, 2).value

            # صف فارغ
            if not ts and not office:
                continue

            parsed = parse_row(ws, r, sheet_type)

            # فحص صفوف الاختبار
            if is_test_row(parsed):
                parsed["exclude_reason"] = "صف اختباري (نمط اسم تجريبي أو مهمة اختبار)"
                excluded.append(parsed)
                continue

            # فحص غياب الطابع الزمني
            if not parsed["timestamp"]:
                parsed["exclude_reason"] = "طابع زمني فارغ"
                excluded.append(parsed)
                continue

            # فحص غياب اسم المكتب
            if not parsed["office_raw"]:
                parsed["exclude_reason"] = "اسم المكتب فارغ"
                excluded.append(parsed)
                continue

            # فحص غياب المهام
            if not parsed["tasks"]:
                parsed["exclude_reason"] = "لا توجد مهام مكتملة في الصف"
                excluded.append(parsed)
                continue

            accepted.append(parsed)

    return accepted, excluded


# ─── فحص التكرارات ────────────────────────────────────────────────────────────

def detect_duplicates(rows: list[dict]) -> list[tuple]:
    """
    يكتشف صفوف متكررة (نفس المكتب + نفس الشهر + نفس السنة).
    يُعيد قائمة من (key, [row_indices]) للمجموعات المكررة.
    """
    seen: dict[tuple, list] = {}
    for i, row in enumerate(rows):
        key = (row["office_raw"], row["month"], row["year"])
        seen.setdefault(key, []).append(i)

    return [(k, v) for k, v in seen.items() if len(v) > 1]


# ─── مطابقة أسماء المكاتب ────────────────────────────────────────────────────

def fetch_offices(db) -> list[dict]:
    """يجلب قائمة المكاتب من Supabase."""
    result = db.table("offices").select("id, name").execute()
    return result.data or []


def build_office_suggestions(
    excel_offices: list[str],
    db_offices: list[dict],
) -> dict[str, dict]:
    """
    يُنشئ قاموس: office_raw → {office_id, office_db_name, similarity}
    لأقرب مطابقة في قاعدة البيانات.
    """
    suggestions = {}
    for raw in excel_offices:
        best_score = -1.0
        best_match = None
        for db_off in db_offices:
            score = similarity(raw, db_off["name"])
            if score > best_score:
                best_score = score
                best_match = db_off
        if best_match:
            suggestions[raw] = {
                "office_id": best_match["id"],
                "office_db_name": best_match["name"],
                "similarity": round(best_score, 3),
            }
        else:
            suggestions[raw] = {
                "office_id": None,
                "office_db_name": None,
                "similarity": 0.0,
            }
    return suggestions


# ─── تقرير Dry-Run ───────────────────────────────────────────────────────────

def print_dry_run_report(
    accepted: list[dict],
    excluded: list[dict],
    suggestions: dict[str, dict],
    duplicates: list[tuple],
) -> None:
    """يطبع تقرير dry-run الكامل."""
    SEP = "─" * 72

    print(f"\n{'═'*72}")
    print("  تقرير DRY-RUN — ترحيل البيانات التاريخية إلى Supabase")
    print(f"{'═'*72}\n")

    # ─── الملخص العام ────────────────────────────────────────────────────────
    total_tasks = sum(len(r["tasks"]) for r in accepted)
    print(f"📊 الملخص الإجمالي:")
    print(f"   الصفوف الكلية (بعد الفلتر الأولي): {len(accepted) + len(excluded)}")
    print(f"   الصفوف المقبولة للترحيل          : {len(accepted)}")
    print(f"   الصفوف المستبعدة                 : {len(excluded)}")
    print(f"   إجمالي submissions متوقعة         : {len(accepted)}")
    print(f"   إجمالي tasks متوقعة               : {total_tasks}")

    # ─── صفوف استُخدم فيها الطابع الزمني لاشتقاق الشهر ──────────────────────
    ts_derived = [r for r in accepted if r["month_source"] == "timestamp"]
    if ts_derived:
        print(f"\n{SEP}")
        print(f"⚠️  صفوف اشتُقّ فيها الشهر/السنة من الطابع الزمني ({len(ts_derived)} صف):")
        print("   (يرجى مراجعتها للتأكد من صحة الشهر)\n")
        for r in ts_derived:
            print(
                f"   [{r['sheet']} | صف {r['row']}] "
                f"المكتب: {r['office_raw']!r:30s} "
                f"| الطابع: {r['timestamp']} "
                f"→ شهر {r['month']}/{r['year']}"
            )

    # ─── الصفوف المستبعدة ────────────────────────────────────────────────────
    if excluded:
        print(f"\n{SEP}")
        print(f"🚫 الصفوف المستبعدة ({len(excluded)}):\n")
        for r in excluded:
            print(
                f"   [{r['sheet']} | صف {r['row']}] "
                f"{r['office_raw']!r:30s} "
                f"→ السبب: {r['exclude_reason']}"
            )

    # ─── فحص التكرارات ───────────────────────────────────────────────────────
    print(f"\n{SEP}")
    if duplicates:
        print(f"🔴 تكرارات مكتشفة ({len(duplicates)}) — السكربت سيتوقف بخطأ:\n")
        for (office, month, year), indices in duplicates:
            rows_info = [f"صف {accepted[i]['row']}" for i in indices]
            print(f"   المكتب: {office!r} | الشهر: {month}/{year} → {', '.join(rows_info)}")
    else:
        print("✅ لا توجد تكرارات (نفس المكتب + نفس الشهر).")

    # ─── مطابقة المكاتب ──────────────────────────────────────────────────────
    print(f"\n{SEP}")
    print("🏢 مطابقة أسماء المكاتب (يرجى مراجعتها وتأكيدها):\n")
    print(f"   {'اسم الإكسل':35s} {'أقرب مطابقة في DB':35s} {'تشابه':8s} {'DB_ID':6s}")
    print(f"   {'-'*35} {'-'*35} {'-'*8} {'-'*6}")
    for raw, sug in suggestions.items():
        sim_pct = f"{sug['similarity']*100:.1f}%"
        db_name = sug['office_db_name'] or "❌ غير موجود"
        db_id = str(sug['office_id']) if sug['office_id'] else "—"
        flag = "⚠️ " if sug['similarity'] < 0.7 else "   "
        print(f"   {flag}{raw:33s} {db_name:35s} {sim_pct:8s} {db_id}")

    print(f"\n{SEP}")
    print("📋 الخطوات التالية:")
    print("   1. راجع مطابقات المكاتب أعلاه")
    print("   2. أنشئ/عدّل ملف office_name_mapping.json بناءً على مراجعتك")
    print("   3. شغّل السكربت بـ --commit للتنفيذ الفعلي")
    print(f"{'═'*72}\n")


# ─── وضع Commit: الإدخال الفعلي ──────────────────────────────────────────────

def load_mapping(mapping_path: Path) -> dict[str, int]:
    """يُحمّل ملف المطابقة: {office_raw_name → office_id}."""
    if not mapping_path.exists():
        print(f"❌ ملف المطابقة غير موجود: {mapping_path}")
        print("   شغّل أولاً وضع dry-run، راجع المطابقات، ثم أنشئ الملف.")
        sys.exit(1)
    with open(mapping_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    # التحقق من البنية
    mapping = {}
    for entry in data.get("offices", []):
        raw = entry.get("excel_name")
        oid = entry.get("office_id")
        if raw and oid:
            mapping[raw] = int(oid)
    return mapping


def insert_submission(db, row: dict, office_id: int) -> int:
    """يُدرج submission واحدة ويُعيد submission_id."""
    result = db.table("submissions").insert({
        "office_id": office_id,
        "submitter_name": row["submitter_name"],
        "submitter_phone": row["submitter_phone"],
        "month": row["month"],
        "year": row["year"],
        "has_plan": bool(row["plan_file"]),
        "plan_file_path": row["plan_file"] or None,
        "general_challenges": row["general_challenges"] or None,
        "additional_notes": row["additional_notes"] or None,
        "status": "pending",
    }).execute()
    return result.data[0]["id"]


def insert_tasks(db, tasks: list[dict], submission_id: int) -> None:
    """يُدرج المهام المرتبطة بـ submission."""
    if not tasks:
        return
    tasks_data = [
        {
            "submission_id": submission_id,
            "task_order": i,
            "manager_name": t["manager_name"] or None,
            "manager_phone": t["manager_phone"] or None,
            "task_name": t["task_name"],
            "task_description": t["task_description"] or None,
            "task_type": t["task_type"] or None,
            "execution_mechanism": t["execution_mechanism"] or None,
            "task_status": t["task_status"] or None,
            "issues": t["issues"] or None,
        }
        for i, t in enumerate(tasks, start=1)
    ]
    db.table("tasks").insert(tasks_data).execute()


def enqueue_submission(db, submission_id: int) -> None:
    """يُدرج صفاً في ai_processing_queue — نفس منطق routes/submissions.py."""
    try:
        db.table("ai_processing_queue").insert({
            "submission_id": submission_id,
            "status": "pending",
        }).execute()
    except Exception as e:
        print(f"   ⚠️  فشل إدراج submission {submission_id} في الطابور: {e}")


def run_commit(
    accepted: list[dict],
    mapping: dict[str, int],
    batch_path: Path,
) -> None:
    """يُنفّذ الإدخال الفعلي بعد التحقق من الـ mapping."""
    # تحقق: كل مكتب في القائمة موجود في الـ mapping
    missing = [r["office_raw"] for r in accepted if r["office_raw"] not in mapping]
    if missing:
        unique_missing = sorted(set(missing))
        print("❌ المكاتب التالية غير موجودة في office_name_mapping.json:")
        for m in unique_missing:
            print(f"   - {m!r}")
        print("   أضفها للملف أو صحّح أسماءها قبل التشغيل.")
        sys.exit(1)

    # إنشاء مجلد الترحيل
    MIGRATION_DIR.mkdir(parents=True, exist_ok=True)

    from database import get_supabase_client
    db = get_supabase_client()

    submitted_ids: list[int] = []
    skipped: list[dict] = []
    errors: list[dict] = []
    run_time = datetime.now(timezone.utc).isoformat()

    print(f"\n🚀 بدء الإدخال الفعلي — {len(accepted)} submission...\n")

    with tqdm(total=len(accepted), unit="submission", desc="الترحيل") as pbar:
        for row in accepted:
            office_id = mapping[row["office_raw"]]
            
            # التحقق من وجود الإرسال مسبقاً (لتجنب التكرار)
            try:
                existing = db.table("submissions").select("id").eq("office_id", office_id).eq("month", row["month"]).eq("year", row["year"]).execute()
                if existing.data:
                    skipped.append({
                        "row": row["row"],
                        "office": row["office_raw"],
                        "reason": "already_exists",
                        "existing_id": existing.data[0]["id"]
                    })
                    continue
            except Exception as e:
                errors.append({
                    "row": row["row"],
                    "sheet": row["sheet"],
                    "office": row["office_raw"],
                    "error": f"خطأ أثناء التحقق من التكرار: {e}",
                })
                tqdm.write(f"   ❌ خطأ في صف {row['row']} ({row['office_raw']}): {e}")
                continue
            finally:
                pbar.update(1)

            sub_id = None
            try:
                sub_id = insert_submission(db, row, office_id)
                insert_tasks(db, row["tasks"], sub_id)
                enqueue_submission(db, sub_id)
                submitted_ids.append(sub_id)
                pbar.set_postfix({"آخر_ID": sub_id, "أخطاء": len(errors), "تخطي": len(skipped)})
            except Exception as e:
                if sub_id:
                    try:
                        # Manual rollback
                        db.table("submissions").delete().eq("id", sub_id).execute()
                        tqdm.write(f"   ⚠️ تم التراجع (حذف submission_id={sub_id}) بسبب خطأ.")
                    except Exception as rb_e:
                        tqdm.write(f"   🚨 فشل التراجع لـ submission_id={sub_id}: {rb_e}")
                errors.append({
                    "row": row["row"],
                    "sheet": row["sheet"],
                    "office": row["office_raw"],
                    "error": str(e),
                })
                tqdm.write(f"   ❌ خطأ في صف {row['row']} ({row['office_raw']}): {e}")

    # ─── ملف batch ───────────────────────────────────────────────────────────
    ts_str = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    batch_file = MIGRATION_DIR / f"batch_{ts_str}.json"
    batch_data = {
        "run_time": run_time,
        "total_inserted": len(submitted_ids),
        "total_skipped": len(skipped),
        "total_errors": len(errors),
        "submission_ids": submitted_ids,
        "skipped": skipped,
        "errors": errors,
    }
    with open(batch_file, "w", encoding="utf-8") as f:
        json.dump(batch_data, f, ensure_ascii=False, indent=2)

    print(f"\n{'═'*72}")
    print(f"✅ اكتمل الترحيل:")
    print(f"   submission تم إدخالها : {len(submitted_ids)}")
    print(f"   تم تخطيها (تكرار)    : {len(skipped)}")
    print(f"   أخطاء                : {len(errors)}")
    print(f"   ملف batch             : {batch_file}")
    print(f"{'═'*72}\n")


# ─── النقطة الرئيسية ──────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="ترحيل بيانات الإكسل التاريخية إلى Supabase."
    )
    parser.add_argument("xlsx_path", help="مسار ملف الإكسل المصدر")
    parser.add_argument(
        "--commit",
        action="store_true",
        help="تنفيذ الإدخال الفعلي (بدونه: dry-run فقط)",
    )
    parser.add_argument(
        "--mapping",
        default=str(DEFAULT_MAPPING_FILE),
        help="مسار ملف office_name_mapping.json (يُستخدم مع --commit)",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.exists():
        print(f"❌ الملف غير موجود: {xlsx_path}")
        sys.exit(1)

    mapping_path = Path(args.mapping)

    # ─── تحميل البيانات ───────────────────────────────────────────────────────
    print(f"📂 تحميل الملف: {xlsx_path}")
    accepted, excluded = load_excel(xlsx_path)
    print(f"   {len(accepted)} صف مقبول، {len(excluded)} مستبعد")

    # ─── فحص التكرارات (في كلا الوضعين) ─────────────────────────────────────
    duplicates = detect_duplicates(accepted)

    if args.commit:
        # ─── وضع Commit ───────────────────────────────────────────────────────
        if duplicates:
            print("\n❌ تكرارات غير محلولة — يُرجى مراجعة dry-run أولاً:")
            for (office, month, year), indices in duplicates:
                rows_info = [f"صف {accepted[i]['row']}" for i in indices]
                print(f"   {office!r} | {month}/{year} → {', '.join(rows_info)}")
            sys.exit(1)

        mapping = load_mapping(mapping_path)
        run_commit(accepted, mapping, MIGRATION_DIR)

    else:
        # ─── وضع Dry-Run ──────────────────────────────────────────────────────
        # جلب المكاتب من Supabase للمطابقة
        try:
            from database import get_supabase_client
            db = get_supabase_client()
            db_offices = fetch_offices(db)
        except Exception as e:
            print(f"\n❌ خطأ حرج: تعذّر الاتصال بقاعدة بيانات Supabase لعمل مطابقة المكاتب: {e}")
            sys.exit(1)

        unique_offices = sorted(set(r["office_raw"] for r in accepted))
        suggestions = build_office_suggestions(unique_offices, db_offices)

        print_dry_run_report(accepted, excluded, suggestions, duplicates)

        # إذا وُجدت تكرارات نُخبر المستخدم أن --commit سيرفض التشغيل
        if duplicates:
            print("⛔ التشغيل بـ --commit سيتوقف بسبب التكرارات أعلاه.")
            print("   يرجى تنظيف الملف وإعادة تشغيل dry-run.\n")
            sys.exit(2)

        # إذا لا يوجد ملف mapping → أنشئ مقترحاً
        if not mapping_path.exists():
            offices_list = [
                {
                    "excel_name": raw,
                    "office_id": sug["office_id"],
                    "office_db_name": sug["office_db_name"],
                    "similarity": sug["similarity"],
                    "_note": "راجع وأكّد أو صحّح office_id قبل --commit",
                }
                for raw, sug in suggestions.items()
            ]
            draft = {"offices": offices_list}
            mapping_path.parent.mkdir(parents=True, exist_ok=True)
            with open(mapping_path, "w", encoding="utf-8") as f:
                json.dump(draft, f, ensure_ascii=False, indent=2)
            print(f"📄 تم إنشاء مسودة mapping في: {mapping_path}")
            print("   راجع الملف، أكّد office_id لكل مكتب، ثم شغّل --commit.\n")


if __name__ == "__main__":
    main()
